#Memuat Module external
import cv2
import face_recognition
import os
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#Membaca data yang ada di path
path = 'D:\\Coding\\Python\\Facedetection\\Student_Face'
images = []
classNames = []
mylist = os.listdir(path)

for cl in mylist:
    curImg = cv2.imread(os.path.join(path, cl))
    if curImg is not None:  # Pastikan gambar dibaca dengan benar
        images.append(curImg)
        classNames.append(os.path.splitext(cl)[0])

#Mengubah data menjadi dapat dibaca komputer
def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode_face = face_recognition.face_encodings(img)
        if encode_face:  # Tambahkan hanya jika encoding berhasil
            encodeList.append(encode_face[0])
    return encodeList

encoded_face_train = findEncodings(images)

#Mengkonversi data ke Excel
def markAttendance(name, wb):
    ws = wb.active 
    ws['B2'] = "Nama"
    ws['C2'] = "Waktu"
    ws['D2'] = "Tanggal"
    start_row = 3
    start_column = 2
    myDataList = []
    
    for row in ws.iter_rows(min_row=start_row, min_col=start_column, values_only=True):
        myDataList.append(row[0])

    if name not in myDataList:
        now = datetime.now()
        time = now.strftime('%I:%M:%S %p')
        date = now.strftime('%d-%B-%Y')

        ws.cell(row=len(myDataList) + start_row, column=start_column, value=name)
        ws.cell(row=len(myDataList) + start_row, column=start_column + 1, value=time)
        ws.cell(row=len(myDataList) + start_row, column=start_column + 2, value=date)
        
        column_widths = [max(len(str(cell.value)) for cell in column) for column in ws.iter_cols(min_col=start_column, max_col=start_column + 2, min_row=start_row, max_row=len(myDataList) + start_row)]
        for i, width in enumerate(column_widths, start=start_column):
            ws.column_dimensions[get_column_letter(i)].width = width * 1.2 
        wb.save(r'D:\Coding\Python\Facedetection\Attendance.xlsx')

# Memuat workbook Excel
attendance_wb = load_workbook(filename=r'D:\Coding\Python\Facedetection\Attendance.xlsx')

#Mengakses dan membaca gambar dari kamera
cap = cv2.VideoCapture(0)
while True:
    success, img = cap.read()
    if not success:
        print("Failed to capture image")
        break
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
    faces_in_frame = face_recognition.face_locations(imgS)
    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)

    #Membandingkan wajah di kamera dengan wajah dari data yang ada
    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
        if encoded_face_train:
            matches = face_recognition.compare_faces(encoded_face_train, encode_face)
            faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
            matchIndex = np.argmin(faceDist) if matches else None

            # Pastikan ada kecocokan dan index sesuai
            if matchIndex is not None and matches[matchIndex]:
                name = classNames[matchIndex].upper()
                y1, x2, y2, x1 = faceloc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                markAttendance(name, attendance_wb)

    #Menampilkan hasil di jendela 'webcam'
    cv2.imshow('webcam', img)
    if cv2.waitKey(1) & 0xFF == ord('e'):
        break

cap.release()
cv2.destroyAllWindows()
attendance_wb.close()  # Menutup workbook setelah selesai
