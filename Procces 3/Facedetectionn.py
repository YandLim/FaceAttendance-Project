import cv2
import face_recognition
import os
import numpy as np
from datetime import datetime
from openpyxl import Workbook

path = 'D:/Coding/Python/Facedetection/Procces 3/Student_Face'

images = []
classNames = []
mylist = os.listdir(path)
for cl in mylist:
    curImg = cv2.imread(os.path.join(path, cl))
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])

def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode_face = face_recognition.face_encodings(img)[0]
        encodeList.append(encode_face)
    return encodeList

encoded_face_train = findEncodings(images)

from openpyxl import load_workbook

def markAttendance(name, wb):
    ws = wb.active
    start_row = 4
    start_column = 2
    myDataList = []
    
    for row in ws.iter_rows(min_row=start_row, min_col=start_column, values_only=True):
        myDataList.append(row[0])

    if name not in myDataList:
        now = datetime.now()
        time = now.strftime('%I:%M:%S:%p')
        date = now.strftime('%d-%B-%Y')

        ws.cell(row=len(myDataList) + start_row, column=start_column, value=name)
        ws.cell(row=len(myDataList) + start_row, column=start_column + 1, value=time)
        ws.cell(row=len(myDataList) + start_row, column=start_column + 2, value=date)
        wb.save(r'D:\Coding\Python\Facedetection\Procces 3\Attendance.xlsx')

attendance_wb = load_workbook(filename=r'D:\Coding\Python\Facedetection\Procces 3\Attendance.xlsx')

cap = cv2.VideoCapture(0)
while True:
    success, img = cap.read()
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
    faces_in_frame = face_recognition.face_locations(imgS)
    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)

    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
        matchIndex = np.argmin(faceDist)

        if matches[matchIndex]:
            name = classNames[matchIndex].upper().lower()
            y1, x2, y2, x1 = faceloc
            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
            markAttendance(name, attendance_wb)

    cv2.imshow('webcam', img)
    if cv2.waitKey(1) & 0xFF == ord('e'):
        break

cap.release()
cv2.destroyAllWindows()